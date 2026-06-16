import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEETS = {
    "characters": [
        "monster_id", "name", "glyph", "base_hp", "base_speed", "slots",
        "base_trait", "description",
    ],
    "character_levels": [
        "monster_id", "level", "hp_add", "speed_add", "trait_id",
    ],
    "traits": ["trait_id", "name", "description"],
    "skills": [
        "skill_id", "name", "glyph", "description", "target", "targets",
        "effect_type",
    ],
    "skill_levels": [
        "skill_id", "level", "power", "ct", "hits", "multiplier",
        "guard_pct", "speed_pct", "duration",
    ],
    "stages": [
        "stage_id", "name", "category", "info", "unlock_type",
        "unlock_value",
    ],
    "stage_waves": [
        "stage_id", "wave", "position", "monster_id", "level",
    ],
    "drops": [
        "source_monster_id", "source_level", "drop_type", "drop_id", "rate",
    ],
    "recipes": ["recipe_id", "name", "result_type", "result_id"],
    "recipe_materials": [
        "recipe_id", "material_monster_id", "count",
    ],
    "system_rules": ["rule_id", "value", "description"],
}


DATA = {
    "characters": [
        ["M_BLACKSQUARE", "Blacksquare", "■", 20, 10, 2, "", "Level3でワイドガードを獲得する基礎個体。"],
        ["M_BLACKCIRCLE", "BlackCircle", "●", 10, 10, 3, "", "Level3で通常攻撃が2体対象になる高速個体。"],
        ["M_HIDER", "hider", "◎", 10, 15, 2, "", "Level3で毎ターン30%の確率で潜伏する。"],
        ["M_BLACKTRIANGLE", "Blacktriangle", "▲", 15, 10, 2, "", "Level3で通常攻撃が同じ敵への3連射になる。"],
        ["M_SPEEDER", "speeder", "▲", 10, 20, 2, "", "Level3でスピードが20%上昇する。"],
        ["M_OUMUAMUA", "オウムアムア", "◉", 30, 50, 3, "", "高速で突入する異形。Level3で光速突進を行う。"],
        ["M_HAMMER", "ハンマー", "T", 10, 5, 1, "", "1ターン溜めて強打を行う低速個体。"],
        ["M_POZ", "ポズ", "♟", 10, 10, 1, "poisonSecretion", "通常攻撃で毒を分泌する状態異常型個体。"],
    ],
    "character_levels": [
        ["M_BLACKSQUARE", 2, 10, 5, ""], ["M_BLACKSQUARE", 3, 20, 5, "wideGuard"],
        ["M_BLACKCIRCLE", 2, 10, 5, ""], ["M_BLACKCIRCLE", 3, 10, 5, "scatter"],
        ["M_HIDER", 2, 10, 10, ""], ["M_HIDER", 3, 10, 15, "hide"],
        ["M_BLACKTRIANGLE", 2, 10, 10, ""], ["M_BLACKTRIANGLE", 3, 20, 5, "rapid"],
        ["M_SPEEDER", 2, 10, 15, ""], ["M_SPEEDER", 3, 10, 15, "dash"],
        ["M_OUMUAMUA", 2, 20, 30, "barrage"], ["M_OUMUAMUA", 3, 20, 40, "lightRush"],
        ["M_HAMMER", 2, 10, 5, ""], ["M_HAMMER", 3, 15, 5, "heavyStrike"],
        ["M_POZ", 2, 6, 15, ""], ["M_POZ", 3, 7, 15, ""],
    ],
    "traits": [
        ["wideGuard", "ワイドガード", "パーティー全体の受けるダメージを軽減。重複は2体まで。"],
        ["scatter", "散弾", "通常攻撃が敵2体を自動対象にする。"],
        ["hide", "ハイド", "毎ターン30%の確率で潜伏する。"],
        ["rapid", "連射", "通常攻撃が選んだ敵へ3ヒットする。"],
        ["dash", "ダッシュ", "このキャラクターのスピードが20%上昇する。"],
        ["barrage", "弾幕", "通常攻撃が敵3体を自動対象にする。"],
        ["lightRush", "光速突進", "ターン最初の行動時、通常攻撃がSの0.5倍ダメージになる。"],
        ["heavyStrike", "強打", "通常攻撃を選ぶと1ターン溜め、次ターンに現在HPと同じダメージを与える。"],
        ["poisonSecretion", "毒分泌", "通常攻撃命中時、30%で対象を2ターン毒状態にする。"],
    ],
    "skills": [
        ["SK_BLACKSQUARE", "ワイドガード", "■", "味方全体の被ダメージを軽減。", "self", 0, "guard"],
        ["SK_BLACKCIRCLE", "BlackCircle", "●", "敵2体へダメージ。", "enemy", 2, "damage"],
        ["SK_HIDE", "ハイド", "◎", "このターン使用者が潜伏。", "self", 0, "hide"],
        ["SK_RAPID", "連射", "▲", "敵1体へ3回ダメージ。", "enemy", 1, "rapid"],
        ["SK_DASH", "ダッシュ", "△", "味方1体のSを3ターン上昇。", "ally", 1, "speed_buff"],
        ["SK_BARRAGE", "弾幕", "⋯", "敵3体へダメージ。", "enemy", 3, "damage"],
        ["SK_OUMUAMUA", "オウムアムア", "◉", "自分のSに応じた光速突進。", "enemy", 1, "speed_strike"],
        ["SK_HEAVY_STRIKE", "強打", "T", "1ターン溜め、次ターンに現在HPと同じダメージ。", "enemy", 1, "hp_strike"],
        ["SK_POISON_SECRETION", "分泌毒", "♟", "通常攻撃後、対象へ毒を確定付与。", "enemy", 1, "poison_attack"],
    ],
    "skill_levels": [
        ["SK_BLACKSQUARE", 1, "", 4, "", "", .2, "", ""],
        ["SK_BLACKSQUARE", 2, "", 3, "", "", .2, "", ""],
        ["SK_BLACKSQUARE", 3, "", 2, "", "", .3, "", ""],
        ["SK_BLACKCIRCLE", 1, 7, 0, "", "", "", "", ""],
        ["SK_BLACKCIRCLE", 2, 10, 0, "", "", "", "", ""],
        ["SK_BLACKCIRCLE", 3, 14, 0, "", "", "", "", ""],
        ["SK_HIDE", 1, "", 5, "", "", "", "", ""],
        ["SK_HIDE", 2, "", 4, "", "", "", "", ""],
        ["SK_HIDE", 3, "", 3, "", "", "", "", ""],
        ["SK_RAPID", 1, "", 6, 3, .7, "", "", ""],
        ["SK_RAPID", 2, "", 5, 3, .7, "", "", ""],
        ["SK_RAPID", 3, "", 4, 3, .8, "", "", ""],
        ["SK_DASH", 1, "", 8, "", "", "", .2, 3],
        ["SK_DASH", 2, "", 7, "", "", "", .2, 3],
        ["SK_DASH", 3, "", 6, "", "", "", .3, 3],
        ["SK_BARRAGE", 1, 8, 7, "", "", "", "", ""],
        ["SK_BARRAGE", 2, 10, 6, "", "", "", "", ""],
        ["SK_BARRAGE", 3, 10, 5, "", "", "", "", ""],
        ["SK_OUMUAMUA", 1, "", 10, "", .5, "", "", ""],
        ["SK_OUMUAMUA", 2, "", 9, "", .5, "", "", ""],
        ["SK_OUMUAMUA", 3, "", 8, "", .7, "", "", ""],
        ["SK_HEAVY_STRIKE", 1, "", 8, 1, 1, "", "", 1],
        ["SK_HEAVY_STRIKE", 2, "", 7, 1, 1, "", "", 1],
        ["SK_HEAVY_STRIKE", 3, "", 6, 1, 1, "", "", 1],
        ["SK_POISON_SECRETION", 1, 5, 6, 1, "", "", "", 2],
        ["SK_POISON_SECRETION", 2, 5, 5, 1, "", "", "", 2],
        ["SK_POISON_SECRETION", 3, 5, 4, 1, "", "", "", 3],
    ],
    "stages": [
        ["STAGE_001", "BlackCircle", "normal", "1 WAVE / BlackCircle x1", "always", ""],
        ["STAGE_002", "Blacksquare", "normal", "2 WAVES / BlackCircle x2 / Blacksquare Lv3", "clear", "STAGE_001"],
        ["STAGE_003", "Blacksquare＆BlackCircle", "normal", "3 WAVES / Lv2-3 MIX", "clear", "STAGE_002"],
        ["STAGE_004", "hide", "normal", "1 WAVE / hider x1", "clear", "STAGE_001"],
        ["STAGE_005", "hide!!!", "normal", "3 WAVES / hider + Blacktriangle", "clear", "STAGE_004"],
        ["STAGE_006", "dash", "normal", "1 WAVE / speeder x1", "clear", "STAGE_001"],
        ["STAGE_007", "triangles", "normal", "3 WAVES / speeder + Blacktriangle", "clear", "STAGE_006"],
        ["SP_OUMUAMUA", "オウムアムア", "special", "3 WAVES / オウムアムア", "kills", "M_SPEEDER:30;M_BLACKCIRCLE:30"],
        ["STAGE_008", "鍛冶", "normal", "1 WAVE / ハンマー Lv1", "clear", "STAGE_004"],
        ["STAGE_009", "工房", "normal", "2 WAVES / ハンマー Lv2-3", "clear", "STAGE_008"],
        ["STAGE_010", "工業廃水", "normal", "1 WAVE / ポズ Lv1", "clear", "STAGE_006"],
        ["STAGE_011", "毒沼", "normal", "3 WAVES / ポズ Lv2-3", "clear", "STAGE_010"],
    ],
    "stage_waves": [
        ["STAGE_001", 1, 1, "M_BLACKCIRCLE", 1],
        ["STAGE_002", 1, 1, "M_BLACKCIRCLE", 1], ["STAGE_002", 1, 2, "M_BLACKCIRCLE", 1],
        ["STAGE_002", 2, 1, "M_BLACKSQUARE", 3],
        ["STAGE_003", 1, 1, "M_BLACKCIRCLE", 2], ["STAGE_003", 1, 2, "M_BLACKCIRCLE", 2], ["STAGE_003", 1, 3, "M_BLACKCIRCLE", 2],
        ["STAGE_003", 2, 1, "M_BLACKSQUARE", 2], ["STAGE_003", 2, 2, "M_BLACKSQUARE", 2], ["STAGE_003", 2, 3, "M_BLACKSQUARE", 2],
        ["STAGE_003", 3, 1, "M_BLACKSQUARE", 3], ["STAGE_003", 3, 2, "M_BLACKSQUARE", 3], ["STAGE_003", 3, 3, "M_BLACKCIRCLE", 3], ["STAGE_003", 3, 4, "M_BLACKCIRCLE", 3],
        ["STAGE_004", 1, 1, "M_HIDER", 1],
        ["STAGE_005", 1, 1, "M_HIDER", 2], ["STAGE_005", 1, 2, "M_HIDER", 2], ["STAGE_005", 1, 3, "M_HIDER", 2],
        ["STAGE_005", 2, 1, "M_BLACKTRIANGLE", 2], ["STAGE_005", 2, 2, "M_BLACKTRIANGLE", 2], ["STAGE_005", 2, 3, "M_BLACKTRIANGLE", 2],
        ["STAGE_005", 3, 1, "M_HIDER", 3], ["STAGE_005", 3, 2, "M_HIDER", 3], ["STAGE_005", 3, 3, "M_BLACKTRIANGLE", 3], ["STAGE_005", 3, 4, "M_BLACKTRIANGLE", 3],
        ["STAGE_006", 1, 1, "M_SPEEDER", 1],
        ["STAGE_007", 1, 1, "M_SPEEDER", 2], ["STAGE_007", 1, 2, "M_BLACKTRIANGLE", 2],
        ["STAGE_007", 2, 1, "M_SPEEDER", 2], ["STAGE_007", 2, 2, "M_SPEEDER", 2], ["STAGE_007", 2, 3, "M_BLACKTRIANGLE", 2], ["STAGE_007", 2, 4, "M_BLACKTRIANGLE", 2],
        ["STAGE_007", 3, 1, "M_SPEEDER", 3], ["STAGE_007", 3, 2, "M_SPEEDER", 3], ["STAGE_007", 3, 3, "M_BLACKTRIANGLE", 3], ["STAGE_007", 3, 4, "M_BLACKTRIANGLE", 3],
        ["SP_OUMUAMUA", 1, 1, "M_OUMUAMUA", 1], ["SP_OUMUAMUA", 1, 2, "M_BLACKCIRCLE", 3], ["SP_OUMUAMUA", 1, 3, "M_BLACKCIRCLE", 3],
        ["SP_OUMUAMUA", 2, 1, "M_OUMUAMUA", 2], ["SP_OUMUAMUA", 2, 2, "M_SPEEDER", 3], ["SP_OUMUAMUA", 2, 3, "M_SPEEDER", 3],
        ["SP_OUMUAMUA", 3, 1, "M_OUMUAMUA", 3],
        ["STAGE_008", 1, 1, "M_HAMMER", 1],
        ["STAGE_009", 1, 1, "M_HAMMER", 2], ["STAGE_009", 1, 2, "M_HAMMER", 2],
        ["STAGE_009", 2, 1, "M_HAMMER", 3], ["STAGE_009", 2, 2, "M_HAMMER", 3], ["STAGE_009", 2, 3, "M_HAMMER", 3],
        ["STAGE_010", 1, 1, "M_POZ", 1],
        ["STAGE_011", 1, 1, "M_POZ", 2], ["STAGE_011", 1, 2, "M_POZ", 2],
        ["STAGE_011", 2, 1, "M_POZ", 3],
        ["STAGE_011", 3, 1, "M_POZ", 3], ["STAGE_011", 3, 2, "M_POZ", 3], ["STAGE_011", 3, 3, "M_POZ", 3],
    ],
    "drops": [
        ["M_BLACKCIRCLE", 1, "monster", "M_BLACKCIRCLE", .2],
        ["M_BLACKCIRCLE", 2, "monster", "M_BLACKCIRCLE", .2],
        ["M_BLACKCIRCLE", 3, "monster", "M_BLACKCIRCLE", .2],
        ["M_BLACKSQUARE", 1, "monster", "M_BLACKSQUARE", .2],
        ["M_BLACKSQUARE", 2, "monster", "M_BLACKSQUARE", .2],
        ["M_BLACKSQUARE", 3, "monster", "M_BLACKSQUARE", .2],
        ["M_BLACKSQUARE", 3, "skill", "SK_BLACKSQUARE", .1],
        ["M_HIDER", 1, "monster", "M_HIDER", .2],
        ["M_HIDER", 2, "monster", "M_HIDER", .2],
        ["M_HIDER", 3, "monster", "M_HIDER", .2],
        ["M_HIDER", 3, "skill", "SK_HIDE", .1],
        ["M_BLACKTRIANGLE", 1, "monster", "M_BLACKTRIANGLE", .2],
        ["M_BLACKTRIANGLE", 2, "monster", "M_BLACKTRIANGLE", .2],
        ["M_BLACKTRIANGLE", 3, "monster", "M_BLACKTRIANGLE", .2],
        ["M_BLACKTRIANGLE", 3, "skill", "SK_RAPID", .1],
        ["M_SPEEDER", 1, "monster", "M_SPEEDER", .2],
        ["M_SPEEDER", 3, "monster", "M_SPEEDER", .2],
        ["M_SPEEDER", 3, "skill", "SK_DASH", .1],
        ["M_OUMUAMUA", 1, "recipe", "R_SK_BARRAGE", .1],
        ["M_OUMUAMUA", 2, "recipe", "R_SK_BARRAGE", .2],
        ["M_OUMUAMUA", 2, "recipe", "R_SK_OUMUAMUA", .1],
        ["M_OUMUAMUA", 3, "recipe", "R_SK_BARRAGE", .3],
        ["M_OUMUAMUA", 3, "recipe", "R_SK_OUMUAMUA", .2],
        ["M_OUMUAMUA", 3, "recipe", "R_M_OUMUAMUA", .2],
        ["M_HAMMER", 1, "monster", "M_HAMMER", .2],
        ["M_HAMMER", 2, "monster", "M_HAMMER", .3],
        ["M_HAMMER", 3, "monster", "M_HAMMER", .3],
        ["M_HAMMER", 3, "skill", "SK_HEAVY_STRIKE", .1],
        ["M_POZ", 1, "monster", "M_POZ", .2],
        ["M_POZ", 2, "monster", "M_POZ", .3],
        ["M_POZ", 3, "monster", "M_POZ", .3],
        ["M_POZ", 3, "skill", "SK_POISON_SECRETION", .1],
    ],
    "recipes": [
        ["R_SK_BARRAGE", "弾幕", "skill", "SK_BARRAGE"],
        ["R_SK_OUMUAMUA", "オウムアムア", "skill", "SK_OUMUAMUA"],
        ["R_M_OUMUAMUA", "オウムアムア", "monster", "M_OUMUAMUA"],
    ],
    "recipe_materials": [
        ["R_SK_BARRAGE", "M_BLACKCIRCLE", 5],
        ["R_SK_OUMUAMUA", "M_BLACKCIRCLE", 5],
        ["R_SK_OUMUAMUA", "M_SPEEDER", 8],
        ["R_M_OUMUAMUA", "M_BLACKCIRCLE", 10],
        ["R_M_OUMUAMUA", "M_SPEEDER", 15],
    ],
    "system_rules": [
        ["NORMAL_DAMAGE", 5, "通常攻撃の基礎ダメージ"],
        ["PARTY_SIZE", 8, "編成可能数"],
        ["FRONT_SIZE", 4, "前衛数"],
        ["MAX_LEVEL", 3, "キャラクター/スキル最大Lv"],
        ["RESERVE_TIMING", "turn_end", "撃破された前衛の交代はターン終了後"],
    ],
}

TEMPLATE_SHEETS = [
    "TEMPLATE_CHARACTER",
    "TEMPLATE_SKILL",
    "TEMPLATE_STAGE",
    "TEMPLATE_DROP",
    "TEMPLATE_RECIPE",
]

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
LABEL_FILL = PatternFill("solid", fgColor="252525")
NOTE_FILL = PatternFill("solid", fgColor="E7E6E6")
THIN_BORDER = Border(
    left=Side(style="thin", color="777777"),
    right=Side(style="thin", color="777777"),
    top=Side(style="thin", color="777777"),
    bottom=Side(style="thin", color="777777"),
)


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="151515")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="E0FF00", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        width = min(48, max(12, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width
    if ws.max_row > 1:
        name = "T_" + "".join(ch for ch in ws.title if ch.isalnum())[:20]
        tab = Table(displayName=name, ref=ws.dimensions)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(tab)


def setup_template(ws, title, description):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor="080808")
    ws["A1"].font = Font(color="E0FF00", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = description
    ws["A2"].fill = NOTE_FILL
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 50


def add_field(ws, row, label, guide="", example="", value=""):
    ws.cell(row, 1, label)
    ws.cell(row, 2, value)
    ws.cell(row, 3, example)
    ws.cell(row, 4, guide)
    ws.cell(row, 1).fill = LABEL_FILL
    ws.cell(row, 1).font = Font(color="FFFFFF", bold=True)
    ws.cell(row, 2).fill = INPUT_FILL
    ws.cell(row, 3).fill = NOTE_FILL
    ws.cell(row, 4).fill = NOTE_FILL
    for cell in ws[row]:
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_list_validation(ws, cell_range, values):
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "一覧から選択してください"
    validation.errorTitle = "入力値エラー"
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_template_sheets(wb):
    for name in TEMPLATE_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet("TEMPLATE_CHARACTER")
    setup_template(ws, "NEW CHARACTER", "黄色セルを埋めます。確定後、characters / character_levels / traits へ転記してください。")
    ws.append(["項目", "入力", "入力例", "説明"])
    fields = [
        ("monster_id", "M_NEWNAME", "M_ + 英大文字。既存IDと重複不可"),
        ("表示名", "NewCharacter", "ゲーム中に表示する名称"),
        ("glyph", "◆", "一覧や簡易描画で使う記号"),
        ("基礎HP Lv1", "20", "Lv1のHP"),
        ("基礎S Lv1", "15", "Lv1のスピード"),
        ("スキル枠", "2", "装備可能なスキル数"),
        ("Lv1特性ID", "", "通常は空欄。trait_idを指定可能"),
        ("説明", "", "キャラクター詳細文"),
        ("Lv2 HP加算", "10", "Lv1からの加算値"),
        ("Lv2 S加算", "5", "Lv1からの加算値"),
        ("Lv2 獲得特性ID", "", "新特性ならtraitsにも追加"),
        ("Lv3 HP加算", "10", "Lv2からの加算値"),
        ("Lv3 S加算", "5", "Lv2からの加算値"),
        ("Lv3 獲得特性ID", "", "新特性ならtraitsにも追加"),
        ("特性名", "", "新特性を追加する場合"),
        ("特性説明", "", "発動条件・効果・重複可否"),
        ("見た目", "", "形状、線、ストライプ、色など"),
        ("専用アニメーション", "", "攻撃・登場・被弾演出"),
    ]
    for row, (label, example, guide) in enumerate(fields, 5):
        add_field(ws, row, label, guide, example)
    add_list_validation(ws, "B10", ["1", "2", "3"])

    ws = wb.create_sheet("TEMPLATE_SKILL")
    setup_template(ws, "NEW SKILL", "黄色セルを埋めます。確定後、skills / skill_levels へ転記してください。")
    ws.append(["項目", "入力", "入力例", "説明"])
    fields = [
        ("skill_id", "SK_NEWNAME", "SK_ + 英大文字。既存IDと重複不可"),
        ("表示名", "NewSkill", "ゲーム中に表示する名称"),
        ("glyph", "◇", "ボタンや一覧で使う記号"),
        ("説明", "", "プレイヤー向け効果文"),
        ("対象", "enemy", "enemy / ally / self"),
        ("対象数", "1", "同時に選ぶ対象数"),
        ("効果タイプ", "damage", "damage / guard / hide / rapid / speed_buff / speed_strike / hp_strike / poison_attack"),
        ("Lv1 Power", "", "固定ダメージ系のみ"),
        ("Lv1 CT", "5", "使用後のクールタイム"),
        ("Lv1 倍率", "", "S倍率・連射倍率など。50%は0.5"),
        ("Lv1 Hits", "", "連射回数"),
        ("Lv1 Guard%", "", "20%は0.2"),
        ("Lv1 Speed%", "", "20%は0.2"),
        ("Lv1 持続ターン", "", "バフなどの持続"),
        ("Lv2変更点", "", "CT、威力、倍率など"),
        ("Lv3変更点", "", "CT、威力、倍率など"),
        ("専用アニメーション", "", "発動演出の指示"),
    ]
    for row, (label, example, guide) in enumerate(fields, 5):
        add_field(ws, row, label, guide, example)
    add_list_validation(ws, "B9", ["enemy", "ally", "self"])
    add_list_validation(
        ws,
        "B11",
        ["damage", "guard", "hide", "rapid", "speed_buff", "speed_strike", "hp_strike", "poison_attack"],
    )

    ws = wb.create_sheet("TEMPLATE_STAGE")
    setup_template(ws, "NEW STAGE", "上部にステージ情報、下部に敵を出現順で入力します。空行は無視してください。")
    ws.append(["項目", "入力", "入力例", "説明"])
    fields = [
        ("stage_id", "STAGE_008", "通常はSTAGE_、テーマはSP_"),
        ("表示名", "new stage", "ステージ選択画面の名称"),
        ("カテゴリ", "normal", "normal / special"),
        ("説明", "", "ステージカードの概要"),
        ("解放条件タイプ", "clear", "always / clear / kills"),
        ("解放条件値", "STAGE_007", "clearならstage_id、killsなら M_ID:30;M_ID:30"),
    ]
    for row, (label, example, guide) in enumerate(fields, 5):
        add_field(ws, row, label, guide, example)
    add_list_validation(ws, "B7", ["normal", "special"])
    add_list_validation(ws, "B9", ["always", "clear", "kills"])
    start = 13
    ws.cell(start, 1, "wave")
    ws.cell(start, 2, "position")
    ws.cell(start, 3, "monster_id")
    ws.cell(start, 4, "level")
    for cell in ws[start]:
        cell.fill = LABEL_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = THIN_BORDER
    for row in range(start + 1, start + 25):
        for col in range(1, 5):
            ws.cell(row, col).fill = INPUT_FILL
            ws.cell(row, col).border = THIN_BORDER
    add_list_validation(ws, f"D{start+1}:D{start+24}", ["1", "2", "3"])

    ws = wb.create_sheet("TEMPLATE_DROP")
    setup_template(ws, "NEW DROPS", "敵Lvごとのドロップ候補を入力します。確率20%は0.2で入力してください。")
    for col, value in enumerate(
        ["source_monster_id", "source_level", "drop_type", "drop_id", "rate"], 1
    ):
        ws.cell(4, col, value)
    ws.column_dimensions["E"].width = 16
    for cell in ws[4]:
        cell.fill = LABEL_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = THIN_BORDER
    for row in range(5, 25):
        for col in range(1, 6):
            ws.cell(row, col).fill = INPUT_FILL
            ws.cell(row, col).border = THIN_BORDER
    add_list_validation(ws, "B5:B24", ["1", "2", "3"])
    add_list_validation(ws, "C5:C24", ["monster", "skill", "recipe"])

    ws = wb.create_sheet("TEMPLATE_RECIPE")
    setup_template(ws, "NEW RECIPE", "レシピ本体と必要素材を入力します。素材は最大10種類まで記入できます。")
    ws.append(["項目", "入力", "入力例", "説明"])
    fields = [
        ("recipe_id", "R_M_NEWNAME", "R_M_キャラ / R_SK_スキル"),
        ("表示名", "NewRecipe", "レシピ一覧の名称"),
        ("結果タイプ", "monster", "monster / skill"),
        ("結果ID", "M_NEWNAME", "獲得するキャラまたはスキルID"),
    ]
    for row, (label, example, guide) in enumerate(fields, 5):
        add_field(ws, row, label, guide, example)
    add_list_validation(ws, "B7", ["monster", "skill"])
    ws.cell(11, 1, "material_monster_id")
    ws.cell(11, 2, "count")
    for cell in ws[11]:
        cell.fill = LABEL_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = THIN_BORDER
    for row in range(12, 22):
        for col in range(1, 3):
            ws.cell(row, col).fill = INPUT_FILL
            ws.cell(row, col).border = THIN_BORDER

    for ws in [wb[name] for name in TEMPLATE_SHEETS]:
        ws.sheet_properties.tabColor = "E0FF00"
        ws.row_dimensions[1].height = 26


def seed_workbook(path):
    if path.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(path, path.with_name(f"{path.stem}_backup_{stamp}{path.suffix}"))
    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    readme.append(["NULL GAME DATABASE", "Excel master -> game_db.json"])
    readme.append(["更新手順", "各DBシートを編集後、tools/export_game_db.py --export-only を実行"])
    readme.append(["IDルール", "既存IDは変更しない。追加時は種別プレフィックスを使用"])
    readme.append(["注意", "stage_waves / drops / recipe_materials は複合行で管理"])
    style_sheet(readme)
    for sheet, headers in SHEETS.items():
        ws = wb.create_sheet(sheet)
        ws.append(headers)
        for row in DATA[sheet]:
            ws.append(row)
        style_sheet(ws)
    add_template_sheets(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def rows(ws):
    headers = [c.value for c in ws[1]]
    return [
        dict(zip(headers, values))
        for values in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in values)
    ]


def level_array(levels, key, default=0):
    result = [default, default, default, default]
    for row in levels:
        value = row.get(key)
        if value not in (None, ""):
            result[int(row["level"])] = value
    return result


def export_json(workbook, output):
    wb = load_workbook(workbook, data_only=True)
    data = {name: rows(wb[name]) for name in SHEETS}
    traits = {
        r["trait_id"]: {"name": r["name"], "desc": r["description"]}
        for r in data["traits"]
    }
    monsters = {}
    for r in data["characters"]:
        levels = {
            str(int(x["level"])): {
                "hp": x["hp_add"] or 0,
                "speed": x["speed_add"] or 0,
                **({"trait": x["trait_id"]} if x["trait_id"] else {}),
            }
            for x in data["character_levels"]
            if x["monster_id"] == r["monster_id"]
        }
        monsters[r["monster_id"]] = {
            "id": r["monster_id"], "name": r["name"], "glyph": r["glyph"],
            "hp": r["base_hp"], "speed": r["base_speed"], "slots": r["slots"],
            "trait": r["base_trait"] or None, "levels": levels,
            "desc": r["description"],
        }
    skills = {}
    for r in data["skills"]:
        lv = [x for x in data["skill_levels"] if x["skill_id"] == r["skill_id"]]
        item = {
            "id": r["skill_id"], "name": r["name"], "glyph": r["glyph"],
            "desc": r["description"], "target": r["target"],
        }
        if r["targets"]:
            item["targets"] = r["targets"]
        effect = r["effect_type"]
        if effect == "damage":
            item["power"] = level_array(lv, "power")
            item["ct"] = level_array(lv, "ct")
        elif effect == "guard":
            item["guardPct"] = level_array(lv, "guard_pct")
            item["ct"] = level_array(lv, "ct")
        elif effect == "hide":
            item["hide"] = True
            item["ct"] = level_array(lv, "ct")
        elif effect == "rapid":
            item["rapid"] = level_array(lv, "multiplier")
            item["hits"] = int(lv[0]["hits"])
            item["ct"] = level_array(lv, "ct")
        elif effect == "speed_buff":
            item["speedPct"] = level_array(lv, "speed_pct")
            item["ct"] = level_array(lv, "ct")
        elif effect == "speed_strike":
            item["speedStrike"] = level_array(lv, "multiplier")
            item["ct"] = level_array(lv, "ct")
        elif effect == "hp_strike":
            item["hpStrike"] = True
            item["ct"] = level_array(lv, "ct")
        elif effect == "poison_attack":
            item["poisonAttack"] = True
            item["poisonTurns"] = level_array(lv, "duration")
            item["ct"] = level_array(lv, "ct")
        skills[r["skill_id"]] = item
    stages = {}
    for r in data["stages"]:
        wave_rows = [x for x in data["stage_waves"] if x["stage_id"] == r["stage_id"]]
        waves = []
        for wave_no in sorted({int(x["wave"]) for x in wave_rows}):
            entries = sorted(
                [x for x in wave_rows if int(x["wave"]) == wave_no],
                key=lambda x: int(x["position"]),
            )
            waves.append([
                {"id": x["monster_id"], "level": int(x["level"])}
                for x in entries
            ])
        stages[r["stage_id"]] = {
            "id": r["stage_id"], "name": r["name"], "category": r["category"],
            "info": r["info"], "unlockType": r["unlock_type"],
            "unlockValue": r["unlock_value"], "waves": waves,
        }
    recipes = {}
    for r in data["recipes"]:
        req = [
            {"id": x["material_monster_id"], "count": int(x["count"])}
            for x in data["recipe_materials"]
            if x["recipe_id"] == r["recipe_id"]
        ]
        recipes[r["recipe_id"]] = {
            "id": r["recipe_id"], "name": r["name"], "kind": r["result_type"],
            "result": {"type": r["result_type"], "id": r["result_id"]},
            "requires": req,
        }
    payload = {
        "schemaVersion": 1,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "monsters": monsters,
        "traits": traits,
        "skills": skills,
        "stages": stages,
        "recipes": recipes,
        "drops": data["drops"],
        "systemRules": {r["rule_id"]: r["value"] for r in data["system_rules"]},
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def update_templates(workbook):
    wb = load_workbook(workbook)
    add_template_sheets(wb)
    wb.save(workbook)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--export-only", action="store_true")
    parser.add_argument("--templates-only", action="store_true")
    args = parser.parse_args()
    if args.templates_only:
        update_templates(args.workbook)
        return
    if not args.output:
        parser.error("--output is required unless --templates-only is used")
    if not args.export_only:
        seed_workbook(args.workbook)
    export_json(args.workbook, args.output)


if __name__ == "__main__":
    main()
